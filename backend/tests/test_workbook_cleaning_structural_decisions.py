import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import duckdb
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from backend.app.workbook_cleaning_apply import apply_cleaning_recipe_to_working_copy
from backend.app.workbook_cleaning_contract import WorkbookCleaningApplyRequest
from backend.app.workbook_cleaning_preview import build_cleaning_recipe_preview


WORKSHEET_ID = "dataset-structural:worksheet:1"


def _evidence(evidence_type, **patch):
    payload = {
        "type": evidence_type,
        "row_index": None,
        "row_range": None,
        "row_indexes": [],
        "column_range": None,
        "label": None,
        "preview_values": [],
        "confidence": "high",
        "explanation": f"{evidence_type} evidence",
    }
    payload.update(patch)
    return payload


def _worksheet_metadata():
    return {
        "worksheet_id": WORKSHEET_ID,
        "sheet_name": "managers",
        "display_name": "managers",
        "original_index": 0,
        "status": "ready",
        "row_count": 6,
        "column_count": 5,
        "schema": [
            {"name": "id"},
            {"name": "name"},
            {"name": "amount"},
            {"name": "note"},
        ],
        "normalization": {
            "header_row_index": 0,
            "template_structure_evidence": [
                _evidence("repeated_header", row_indexes=[3], row_range=[3, 3]),
                _evidence("sparse_layout_gap", row_range=[2, 2]),
                _evidence("serial_only_placeholder_rows", row_range=[4, 4]),
                _evidence("side_note_region_candidate", row_index=0, column_range=[4, 4]),
                _evidence("repeated_missing_pattern", row_indexes=[1], row_range=[1, 1]),
            ],
        },
    }


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "managers"
    rows = [
        ["id", "name", "amount", None, "note"],
        [1, None, 10, None, "n1"],
        [None, None, None, None, None],
        ["id", "name", "amount", None, "note"],
        [99, None, None, None, "placeholder note"],
        [2, "Beta", 20, None, "n2"],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _raw_rows():
    return [
        ["id", "name", "amount", None, "note"],
        [1, None, 10, None, "n1"],
        [None, None, None, None, None],
        ["id", "name", "amount", None, "note"],
        [99, None, None, None, "placeholder note"],
        [2, "Beta", 20, None, "n2"],
    ]


def _decision(evidence_type, decision, index, **patch):
    recommendation_id = f"{WORKSHEET_ID}:{evidence_type}:{index}"
    payload = {
        "recommendation_id": recommendation_id,
        "evidence_type": evidence_type,
        "decision": decision,
        "evidence_ids": [recommendation_id],
    }
    payload.update(patch)
    return payload


def _plan(decisions):
    return WorkbookCleaningApplyRequest(
        structural_decision_plan={
            "worksheet_id": WORKSHEET_ID,
            "decisions": decisions,
        }
    ).structural_decision_plan


class WorkbookCleaningStructuralDecisionRecipeTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "structural.xlsx"
        self.duckdb_path = Path(self.temp_dir.name) / "session.duckdb"
        _write_workbook(self.workbook_path)
        self.parse_patcher = patch(
            "backend.app.workbook_cleaning_preview.parse_xlsx_workbook",
            return_value=[{"name": "managers", "rows": _raw_rows()}],
        )
        self.parse_patcher.start()
        with duckdb.connect(str(self.duckdb_path)) as connection:
            connection.execute("CREATE TABLE source_marker (id VARCHAR)")
            connection.execute("INSERT INTO source_marker VALUES ('original')")
        self.worksheet = _worksheet_metadata()

    def tearDown(self) -> None:
        self.parse_patcher.stop()
        self.temp_dir.cleanup()

    def _preview_rows(self, plan):
        preview = build_cleaning_recipe_preview(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            row_limit=20,
            structural_decision_plan=plan,
        )
        return preview, preview["after_preview"]["row_provenance"]

    def test_accepted_automatic_blank_rows_are_excluded(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "use_recommendation", 0, affected_rows=[2]),
        ])

        preview, provenance = self._preview_rows(plan)

        self.assertNotIn(2, [row["original_row_index"] for row in provenance])
        self.assertEqual(preview["excluded_details"]["layout_rows"]["row_indexes"], [2])

    def test_keep_original_automatic_blank_rows_are_preserved(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "keep_original", 0, affected_rows=[2]),
        ])

        preview, provenance = self._preview_rows(plan)

        self.assertIn(2, [row["original_row_index"] for row in provenance])
        self.assertEqual(preview["excluded"]["layout_rows"], 0)

    def test_deferred_automatic_blank_rows_are_preserved(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "decide_later", 0, affected_rows=[2]),
        ])

        _, provenance = self._preview_rows(plan)

        self.assertIn(2, [row["original_row_index"] for row in provenance])

    def test_accepted_sparse_layout_rows_are_excluded(self) -> None:
        plan = _plan([
            _decision("sparse_layout_gap", "use_recommendation", 1, affected_rows=[2]),
        ])

        _, provenance = self._preview_rows(plan)

        self.assertNotIn(2, [row["original_row_index"] for row in provenance])

    def test_keep_original_layout_rows_are_preserved(self) -> None:
        plan = _plan([
            _decision("sparse_layout_gap", "keep_original", 1, affected_rows=[2]),
        ])

        _, provenance = self._preview_rows(plan)

        self.assertIn(2, [row["original_row_index"] for row in provenance])

    def test_accepted_repeated_headers_are_removed(self) -> None:
        plan = _plan([
            _decision("repeated_header", "use_recommendation", 0, affected_rows=[3]),
        ])

        _, provenance = self._preview_rows(plan)

        self.assertNotIn(3, [row["original_row_index"] for row in provenance])

    def test_keep_original_repeated_headers_are_preserved(self) -> None:
        plan = _plan([
            _decision("repeated_header", "keep_original", 0, affected_rows=[3]),
        ])

        _, provenance = self._preview_rows(plan)

        self.assertIn(3, [row["original_row_index"] for row in provenance])

    def test_accepted_side_note_columns_are_excluded(self) -> None:
        plan = _plan([
            _decision("side_note_region_candidate", "use_recommendation", 3, affected_column_indexes=[4]),
        ])

        preview, _ = self._preview_rows(plan)

        self.assertNotIn("note", preview["after_preview"]["columns"])
        self.assertEqual(preview["excluded"]["side_note_columns"], 1)

    def test_keep_original_side_note_columns_are_preserved(self) -> None:
        plan = _plan([
            _decision("side_note_region_candidate", "keep_original", 3, affected_column_indexes=[4]),
        ])

        preview, _ = self._preview_rows(plan)

        self.assertIn("note", preview["after_preview"]["columns"])
        self.assertEqual(preview["after_preview"]["rows"][0]["note"], "n1")

    def test_deferred_side_note_columns_are_preserved(self) -> None:
        plan = _plan([
            _decision("side_note_region_candidate", "decide_later", 3, affected_column_indexes=[4]),
        ])

        preview, _ = self._preview_rows(plan)

        self.assertIn("note", preview["after_preview"]["columns"])

    def test_repeated_missing_pattern_does_not_fill_values(self) -> None:
        plan = _plan([
            _decision("repeated_missing_pattern", "use_recommendation", 4, affected_rows=[1]),
        ])

        preview, _ = self._preview_rows(plan)

        self.assertIsNone(preview["after_preview"]["rows"][0]["name"])
        self.assertEqual(preview["recipe"][0]["type"], "review_blank_cells")

    def test_mixed_decisions_and_preview_apply_parity(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "use_recommendation", 0, affected_rows=[2]),
            _decision("repeated_header", "keep_original", 0, affected_rows=[3]),
            _decision("side_note_region_candidate", "decide_later", 3, affected_column_indexes=[4]),
        ])

        preview, provenance = self._preview_rows(plan)
        result = apply_cleaning_recipe_to_working_copy(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            duckdb_path=self.duckdb_path,
            dataset_id="dataset-structural",
            structural_decision_plan=plan,
        )

        self.assertNotIn(2, [row["original_row_index"] for row in provenance])
        self.assertIn(3, [row["original_row_index"] for row in provenance])
        self.assertEqual(result["after"]["row_count"], preview["after_preview"]["row_count"])
        self.assertEqual(result["after"]["columns"], preview["after_preview"]["columns"])

    def test_selected_worksheet_only_and_original_workbook_are_unchanged(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "use_recommendation", 0, affected_rows=[2]),
        ])

        result = apply_cleaning_recipe_to_working_copy(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            duckdb_path=self.duckdb_path,
            dataset_id="dataset-structural",
            structural_decision_plan=plan,
        )

        workbook = load_workbook(self.workbook_path)
        self.assertEqual(workbook["managers"].max_row, 6)
        with duckdb.connect(str(self.duckdb_path)) as connection:
            self.assertEqual(connection.execute("SELECT id FROM source_marker").fetchone()[0], "original")
            self.assertEqual(
                connection.execute(f"SELECT COUNT(*) FROM \"{result['cleaned_table_name']}\"").fetchone()[0],
                result["after"]["row_count"],
            )

    def test_invalid_evidence_id_is_rejected_without_partial_write(self) -> None:
        plan = _plan([
            {
                "recommendation_id": f"{WORKSHEET_ID}:automatic_blank_row:0",
                "evidence_type": "automatic_blank_row",
                "decision": "use_recommendation",
                "evidence_ids": [f"{WORKSHEET_ID}:automatic_blank_row:missing"],
                "affected_rows": [2],
            }
        ])

        with self.assertRaises(HTTPException):
            apply_cleaning_recipe_to_working_copy(
                workbook_path=self.workbook_path,
                worksheet=self.worksheet,
                duckdb_path=self.duckdb_path,
                dataset_id="dataset-structural",
                structural_decision_plan=plan,
            )
        with duckdb.connect(str(self.duckdb_path)) as connection:
            tables = {row[0] for row in connection.execute("SHOW TABLES").fetchall()}
        self.assertFalse(any(table.startswith("cleaned_") for table in tables))

    def test_invalid_row_and_column_indexes_are_rejected(self) -> None:
        invalid_row_plan = _plan([
            _decision("automatic_blank_row", "use_recommendation", 0, affected_rows=[99]),
        ])
        invalid_column_plan = _plan([
            _decision("side_note_region_candidate", "use_recommendation", 3, affected_column_indexes=[99]),
        ])

        with self.assertRaises(HTTPException):
            self._preview_rows(invalid_row_plan)
        with self.assertRaises(HTTPException):
            self._preview_rows(invalid_column_plan)

    def test_idempotent_reapply_replaces_same_cleaned_table(self) -> None:
        plan = _plan([
            _decision("automatic_blank_row", "use_recommendation", 0, affected_rows=[2]),
        ])

        first = apply_cleaning_recipe_to_working_copy(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            duckdb_path=self.duckdb_path,
            dataset_id="dataset-structural",
            structural_decision_plan=plan,
        )
        second = apply_cleaning_recipe_to_working_copy(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            duckdb_path=self.duckdb_path,
            dataset_id="dataset-structural",
            structural_decision_plan=plan,
        )

        self.assertEqual(first["cleaned_table_name"], second["cleaned_table_name"])
        self.assertEqual(first["after"]["row_count"], second["after"]["row_count"])


if __name__ == "__main__":
    unittest.main()
