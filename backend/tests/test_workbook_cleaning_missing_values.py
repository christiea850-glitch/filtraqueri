import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import duckdb
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from backend.app.workbook_cleaning_apply import apply_cleaning_recipe_to_working_copy
from backend.app.workbook_cleaning_contract import WorkbookCleaningApplyRequest


WORKSHEET_ID = "dataset-missing:worksheet:1"
OTHER_WORKSHEET_ID = "dataset-missing:worksheet:2"


def _worksheet_metadata():
    return {
        "worksheet_id": WORKSHEET_ID,
        "sheet_name": "Data",
        "display_name": "Data",
        "original_index": 0,
        "status": "ready",
        "row_count": 4,
        "column_count": 6,
        "schema": [
            {"name": "id", "inferred_type": "numeric"},
            {"name": "amount", "inferred_type": "numeric"},
            {"name": "region", "inferred_type": "categorical"},
            {"name": "comment", "inferred_type": "text"},
            {"name": "visit_date", "inferred_type": "date"},
            {"name": "flag", "inferred_type": "boolean"},
        ],
        "normalization": {
            "header_row_index": 0,
            "template_structure_evidence": [],
        },
    }


def _structural_worksheet_metadata():
    metadata = _worksheet_metadata()
    metadata["row_count"] = 5
    metadata["normalization"] = {
        "header_row_index": 0,
        "template_structure_evidence": [
            {
                "type": "automatic_blank_row",
                "row_index": None,
                "row_range": None,
                "row_indexes": [2],
                "column_range": None,
                "label": None,
                "preview_values": [],
                "confidence": "high",
                "explanation": "blank layout row",
            }
        ],
    }
    return metadata


def _rows():
    return [
        ["id", "amount", "region", "comment", "visit_date", "flag"],
        [1, 10, "North", "ready", "2024-01-01", "true"],
        [2, None, None, None, None, "true"],
        [3, 20, "North", "done", "2024-01-03", "false"],
    ]


def _structural_rows():
    return [
        ["id", "amount", "region", "comment", "visit_date", "flag"],
        [1, 10, "North", "ready", "2024-01-01", "true"],
        [None, None, None, None, None, None],
        [2, None, None, None, None, "true"],
        [3, 20, "North", "done", "2024-01-03", "false"],
    ]


def _write_workbook(path: Path, rows) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in rows:
        sheet.append(row)
    other = workbook.create_sheet("Other")
    other.append(["id", "value"])
    other.append([1, "unchanged"])
    workbook.save(path)


def _missing_plan(worksheet_strategy="decide_per_column", column_decisions=None):
    return WorkbookCleaningApplyRequest(
        missing_value_plan={
            "worksheet_id": WORKSHEET_ID,
            "worksheet_strategy": worksheet_strategy,
            "column_decisions": column_decisions or [],
        }
    ).missing_value_plan


def _structural_plan():
    return WorkbookCleaningApplyRequest(
        structural_decision_plan={
            "worksheet_id": WORKSHEET_ID,
            "decisions": [
                {
                    "recommendation_id": f"{WORKSHEET_ID}:automatic_blank_row:0",
                    "evidence_type": "automatic_blank_row",
                    "decision": "use_recommendation",
                    "evidence_ids": [f"{WORKSHEET_ID}:automatic_blank_row:0"],
                    "affected_rows": [2],
                }
            ],
        }
    ).structural_decision_plan


class WorkbookCleaningMissingValueApplyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "missing.xlsx"
        self.duckdb_path = Path(self.temp_dir.name) / "session.duckdb"
        _write_workbook(self.workbook_path, _rows())
        self.parse_patcher = patch(
            "backend.app.workbook_cleaning_preview.parse_xlsx_workbook",
            return_value=[{"name": "Data", "rows": _rows()}],
        )
        self.parse_patcher.start()
        with duckdb.connect(str(self.duckdb_path)) as connection:
            connection.execute("CREATE TABLE original_source (id VARCHAR, amount VARCHAR)")
            connection.execute("INSERT INTO original_source VALUES ('2', NULL)")
            connection.execute("CREATE TABLE unrelated_worksheet (id VARCHAR, value VARCHAR)")
            connection.execute("INSERT INTO unrelated_worksheet VALUES ('1', 'unchanged')")
            connection.execute("CREATE VIEW data AS SELECT * FROM original_source")
        self.worksheet = _worksheet_metadata()

    def tearDown(self) -> None:
        self.parse_patcher.stop()
        self.temp_dir.cleanup()

    def _apply(self, missing_value_plan, structural_decision_plan=None):
        return apply_cleaning_recipe_to_working_copy(
            workbook_path=self.workbook_path,
            worksheet=self.worksheet,
            duckdb_path=self.duckdb_path,
            dataset_id="dataset-missing",
            structural_decision_plan=structural_decision_plan,
            missing_value_plan=missing_value_plan,
        )

    def _cleaned_rows(self, table_name):
        with duckdb.connect(str(self.duckdb_path)) as connection:
            result = connection.execute(f'SELECT * FROM "{table_name}" ORDER BY id')
            columns = [description[0] for description in result.description]
            return [dict(zip(columns, row)) for row in result.fetchall()]

    def test_missing_value_only_creates_cleaned_copy(self) -> None:
        result = self._apply(_missing_plan(column_decisions=[{"column_name": "amount", "strategy": "fill_zero"}]))

        self.assertEqual(result["status"], "applied_to_working_copy")
        self.assertEqual(result["missing_value_summary"]["columns_changed"], ["amount"])
        self.assertEqual(self._cleaned_rows(result["cleaned_table_name"])[1]["amount"], "0")

    def test_combined_structural_then_missing_value_ordering(self) -> None:
        self.parse_patcher.stop()
        self.parse_patcher = patch(
            "backend.app.workbook_cleaning_preview.parse_xlsx_workbook",
            return_value=[{"name": "Data", "rows": _structural_rows()}],
        )
        self.parse_patcher.start()
        self.worksheet = _structural_worksheet_metadata()

        result = self._apply(
            _missing_plan(column_decisions=[{"column_name": "amount", "strategy": "fill_mean"}]),
            structural_decision_plan=_structural_plan(),
        )

        rows = self._cleaned_rows(result["cleaned_table_name"])
        self.assertEqual(result["excluded"]["layout_rows"], 1)
        self.assertEqual(rows[1]["amount"], "15.0")
        self.assertEqual(result["after"]["row_count"], 3)

    def test_fill_mean_and_median_are_computed_from_fresh_rows(self) -> None:
        result = self._apply(
            _missing_plan(column_decisions=[{"column_name": "amount", "strategy": "fill_median"}])
        )
        first_rows = self._cleaned_rows(result["cleaned_table_name"])
        second = self._apply(
            _missing_plan(column_decisions=[{"column_name": "amount", "strategy": "fill_median"}])
        )
        second_rows = self._cleaned_rows(second["cleaned_table_name"])

        self.assertEqual(first_rows[1]["amount"], "15.0")
        self.assertEqual(second_rows[1]["amount"], "15.0")
        self.assertEqual(result["cleaned_table_name"], second["cleaned_table_name"])

    def test_fill_mode_and_mark_unknown(self) -> None:
        result = self._apply(
            _missing_plan(
                column_decisions=[
                    {"column_name": "region", "strategy": "fill_mode"},
                    {"column_name": "comment", "strategy": "mark_unknown"},
                ]
            )
        )
        rows = self._cleaned_rows(result["cleaned_table_name"])

        self.assertEqual(rows[1]["region"], "North")
        self.assertEqual(rows[1]["comment"], "Unknown")

    def test_custom_numeric_text_and_date_values(self) -> None:
        result = self._apply(
            _missing_plan(
                column_decisions=[
                    {"column_name": "amount", "strategy": "fill_custom", "custom_value": 42},
                    {"column_name": "comment", "strategy": "fill_custom", "custom_value": "TBD"},
                    {"column_name": "visit_date", "strategy": "custom_date", "custom_value": "2024-02-01"},
                ]
            )
        )
        rows = self._cleaned_rows(result["cleaned_table_name"])

        self.assertEqual(rows[1]["amount"], "42")
        self.assertEqual(rows[1]["comment"], "TBD")
        self.assertEqual(rows[1]["visit_date"], "2024-02-01")

    def test_row_level_remove_mostly_blank_rows(self) -> None:
        result = self._apply(_missing_plan(worksheet_strategy="remove_mostly_blank_rows"))

        self.assertEqual(result["missing_value_summary"]["rows_removed"], 1)
        self.assertEqual(result["after"]["row_count"], 2)

    def test_no_op_does_not_create_cleaned_copy(self) -> None:
        result = self._apply(_missing_plan(worksheet_strategy="leave_unchanged"))

        self.assertEqual(result["status"], "no_recipe_needed")
        self.assertIsNone(result["cleaned_table_name"])
        with duckdb.connect(str(self.duckdb_path)) as connection:
            tables = {row[0] for row in connection.execute("SHOW TABLES").fetchall()}
        self.assertFalse(any(table.startswith("cleaned_") for table in tables))

    def test_selected_worksheet_only_original_and_active_source_unchanged(self) -> None:
        result = self._apply(_missing_plan(column_decisions=[{"column_name": "amount", "strategy": "fill_zero"}]))

        workbook = load_workbook(self.workbook_path)
        self.assertEqual(workbook["Data"].max_row, 4)
        with duckdb.connect(str(self.duckdb_path)) as connection:
            self.assertIsNone(connection.execute("SELECT amount FROM original_source WHERE id = '2'").fetchone()[0])
            self.assertEqual(connection.execute("SELECT value FROM unrelated_worksheet").fetchone()[0], "unchanged")
            self.assertEqual(connection.execute("SELECT amount FROM data WHERE id = '2'").fetchone()[0], None)
            self.assertEqual(
                connection.execute(f'SELECT amount FROM "{result["cleaned_table_name"]}" WHERE id = ? ', ["2"]).fetchone()[0],
                "0",
            )

    def test_unknown_column_is_rejected_without_partial_write(self) -> None:
        with self.assertRaises(HTTPException):
            self._apply(_missing_plan(column_decisions=[{"column_name": "missing", "strategy": "fill_zero"}]))
        with duckdb.connect(str(self.duckdb_path)) as connection:
            tables = {row[0] for row in connection.execute("SHOW TABLES").fetchall()}
        self.assertFalse(any(table.startswith("cleaned_") for table in tables))

    def test_invalid_strategy_type_combinations_are_rejected(self) -> None:
        cases = [
            {"column_name": "region", "strategy": "fill_zero"},
            {"column_name": "amount", "strategy": "mark_unknown"},
            {"column_name": "amount", "strategy": "fill_mode"},
            {"column_name": "comment", "strategy": "fill_mean"},
            {"column_name": "amount", "strategy": "custom_date", "custom_value": "2024-02-01"},
            {"column_name": "visit_date", "strategy": "fill_custom", "custom_value": "2024-02-01"},
        ]
        for decision in cases:
            with self.subTest(decision=decision):
                with self.assertRaises(HTTPException):
                    self._apply(_missing_plan(column_decisions=[decision]))

    def test_invalid_custom_values_are_rejected(self) -> None:
        cases = [
            {"column_name": "amount", "strategy": "fill_custom"},
            {"column_name": "amount", "strategy": "fill_custom", "custom_value": "not-a-number"},
            {"column_name": "visit_date", "strategy": "custom_date", "custom_value": "not-a-date"},
        ]
        for decision in cases:
            with self.subTest(decision=decision):
                with self.assertRaises(HTTPException):
                    self._apply(_missing_plan(column_decisions=[decision]))

    def test_mean_median_and_mode_without_source_values_are_rejected(self) -> None:
        rows = [
            ["id", "amount", "region", "comment", "visit_date", "flag"],
            [1, None, None, None, None, None],
        ]
        self.parse_patcher.stop()
        self.parse_patcher = patch(
            "backend.app.workbook_cleaning_preview.parse_xlsx_workbook",
            return_value=[{"name": "Data", "rows": rows}],
        )
        self.parse_patcher.start()

        for decision in [
            {"column_name": "amount", "strategy": "fill_mean"},
            {"column_name": "amount", "strategy": "fill_median"},
            {"column_name": "region", "strategy": "fill_mode"},
        ]:
            with self.subTest(decision=decision):
                with self.assertRaises(HTTPException):
                    self._apply(_missing_plan(column_decisions=[decision]))


if __name__ == "__main__":
    unittest.main()
