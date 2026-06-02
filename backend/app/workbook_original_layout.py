from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any
import re
import zipfile
import xml.etree.ElementTree as ET

from fastapi import HTTPException


MAX_ORIGINAL_LAYOUT_ROWS = 200
MAX_ORIGINAL_LAYOUT_COLUMNS = 50
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg_rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _color_value(color: Any) -> str | None:
    if color is None or getattr(color, "type", None) != "rgb":
        return None

    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return None

    normalized = rgb[-6:].upper()
    if len(normalized) != 6 or any(character not in "0123456789ABCDEF" for character in normalized):
        return None

    return f"#{normalized}"


def _border_side(side: Any) -> dict[str, Any]:
    return {
        "style": getattr(side, "style", None),
        "color": _color_value(getattr(side, "color", None)),
    }


def _cell_style(cell: Any) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    return {
        "fill_color": _color_value(fill.fgColor) if fill.fill_type else None,
        "font": {
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "size": font.sz,
            "color": _color_value(font.color),
        },
        "border": {
            "top": _border_side(border.top),
            "right": _border_side(border.right),
            "bottom": _border_side(border.bottom),
            "left": _border_side(border.left),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": bool(alignment.wrap_text),
        },
    }


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def _xml_color(element: ET.Element | None) -> str | None:
    if element is None:
        return None
    rgb = element.attrib.get("rgb")
    if not rgb:
        return None
    normalized = rgb[-6:].upper()
    if len(normalized) != 6 or any(character not in "0123456789ABCDEF" for character in normalized):
        return None
    return f"#{normalized}"


def _xml_border_side(element: ET.Element | None) -> dict[str, Any]:
    return {
        "style": element.attrib.get("style") if element is not None else None,
        "color": _xml_color(element.find("main:color", XML_NS)) if element is not None else None,
    }


def _empty_style() -> dict[str, Any]:
    return {
        "fill_color": None,
        "font": {"bold": False, "italic": False, "size": None, "color": None},
        "border": {
            "top": {"style": None, "color": None},
            "right": {"style": None, "color": None},
            "bottom": {"style": None, "color": None},
            "left": {"style": None, "color": None},
        },
        "alignment": {"horizontal": None, "vertical": None, "wrap_text": False},
    }


def _parse_xml_styles(workbook_zip: zipfile.ZipFile) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(workbook_zip.read("xl/styles.xml"))
    except KeyError:
        return [_empty_style()]

    fonts = root.findall("main:fonts/main:font", XML_NS)
    fills = root.findall("main:fills/main:fill", XML_NS)
    borders = root.findall("main:borders/main:border", XML_NS)
    styles = []
    for cell_format in root.findall("main:cellXfs/main:xf", XML_NS):
        font_id = int(cell_format.attrib.get("fontId", "0"))
        fill_id = int(cell_format.attrib.get("fillId", "0"))
        border_id = int(cell_format.attrib.get("borderId", "0"))
        font = fonts[font_id] if font_id < len(fonts) else None
        fill = fills[fill_id] if fill_id < len(fills) else None
        border = borders[border_id] if border_id < len(borders) else None
        alignment = cell_format.find("main:alignment", XML_NS)
        pattern_fill = fill.find("main:patternFill", XML_NS) if fill is not None else None
        styles.append(
            {
                "fill_color": (
                    _xml_color(pattern_fill.find("main:fgColor", XML_NS))
                    if pattern_fill is not None
                    else None
                ),
                "font": {
                    "bold": font.find("main:b", XML_NS) is not None if font is not None else False,
                    "italic": font.find("main:i", XML_NS) is not None if font is not None else False,
                    "size": (
                        float(font.find("main:sz", XML_NS).attrib["val"])
                        if font is not None and font.find("main:sz", XML_NS) is not None
                        else None
                    ),
                    "color": _xml_color(font.find("main:color", XML_NS)) if font is not None else None,
                },
                "border": {
                    "top": _xml_border_side(border.find("main:top", XML_NS) if border is not None else None),
                    "right": _xml_border_side(border.find("main:right", XML_NS) if border is not None else None),
                    "bottom": _xml_border_side(border.find("main:bottom", XML_NS) if border is not None else None),
                    "left": _xml_border_side(border.find("main:left", XML_NS) if border is not None else None),
                },
                "alignment": {
                    "horizontal": alignment.attrib.get("horizontal") if alignment is not None else None,
                    "vertical": alignment.attrib.get("vertical") if alignment is not None else None,
                    "wrap_text": (
                        alignment.attrib.get("wrapText") in ("1", "true")
                        if alignment is not None
                        else False
                    ),
                },
            }
        )

    return styles or [_empty_style()]


def _cell_indexes(reference: str) -> tuple[int, int]:
    match = re.match(r"([A-Za-z]+)(\d+)", reference)
    if not match:
        return 1, 1

    letters, row_number = match.groups()
    column = 0
    for letter in letters.upper():
        column = column * 26 + ord(letter) - ord("A") + 1
    return int(row_number), column


def _read_shared_strings(workbook_zip: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return [
        "".join(text.text or "" for text in item.findall(".//main:t", XML_NS))
        for item in root.findall("main:si", XML_NS)
    ]


def _read_sheet_paths(workbook_zip: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    relationships_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    relationships = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships_root.findall("pkg_rel:Relationship", XML_NS)
    }
    sheets = []
    for sheet in workbook_root.findall("main:sheets/main:sheet", XML_NS):
        relationship_id = sheet.attrib.get(f"{{{XML_NS['rel']}}}id", "")
        target = relationships.get(relationship_id)
        if not target:
            continue
        sheet_path = f"xl/{target.lstrip('/')}"
        if not sheet_path.startswith("xl/worksheets/") and "worksheets/" in sheet_path:
            sheet_path = "xl/" + sheet_path.split("xl/", 1)[-1]
        sheets.append((sheet.attrib.get("name", "Worksheet"), sheet_path))
    return sheets


def _extract_layout_from_ooxml(
    *,
    workbook_path: Path,
    worksheet_index: int,
    worksheet_id: str,
    row_start: int,
    row_limit: int,
    column_start: int,
    column_limit: int,
) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(workbook_path) as workbook_zip:
            sheets = _read_sheet_paths(workbook_zip)
            if worksheet_index < 0 or worksheet_index >= len(sheets):
                raise HTTPException(status_code=404, detail="Original workbook worksheet is missing")

            worksheet_name, sheet_path = sheets[worksheet_index]
            root = ET.fromstring(workbook_zip.read(sheet_path))
            styles = _parse_xml_styles(workbook_zip)
            shared_strings = _read_shared_strings(workbook_zip)
    except (zipfile.BadZipFile, KeyError, ET.ParseError, ValueError) as error:
        raise HTTPException(status_code=400, detail="Original XLSX workbook could not be read") from error

    dimensions = root.find("main:dimension", XML_NS)
    dimension_reference = dimensions.attrib.get("ref", "A1") if dimensions is not None else "A1"
    last_reference = dimension_reference.split(":")[-1]
    total_rows, total_columns = _cell_indexes(last_reference)
    total_rows = max(total_rows, 1)
    total_columns = max(total_columns, 1)
    row_end = min(total_rows, row_start + row_limit - 1)
    column_end = min(total_columns, column_start + column_limit - 1)

    cells = []
    has_visible_content = False
    for cell in root.findall(".//main:sheetData/main:row/main:c", XML_NS):
        coordinate = cell.attrib.get("r", "A1")
        row, column = _cell_indexes(coordinate)
        if row < row_start or row > row_end or column < column_start or column > column_end:
            continue
        cell_type = cell.attrib.get("t")
        formula = cell.find("main:f", XML_NS)
        value_node = cell.find("main:v", XML_NS)
        inline_text = cell.find("main:is", XML_NS)
        display_value = ""
        if formula is not None:
            display_value = f"={formula.text or ''}"
        elif cell_type == "inlineStr" and inline_text is not None:
            display_value = "".join(text.text or "" for text in inline_text.findall(".//main:t", XML_NS))
        elif value_node is not None:
            display_value = value_node.text or ""
            if cell_type == "s" and display_value.isdigit():
                string_index = int(display_value)
                if string_index < len(shared_strings):
                    display_value = shared_strings[string_index]
            elif cell_type == "b":
                display_value = "TRUE" if display_value == "1" else "FALSE"
        style_index = int(cell.attrib.get("s", "0"))
        style = styles[style_index] if style_index < len(styles) else _empty_style()
        if display_value or style_index:
            cells.append(
                {
                    "row": row,
                    "column": column,
                    "coordinate": coordinate,
                    "display_value": display_value,
                    "is_formula": formula is not None,
                    "style": style,
                }
            )
        if display_value:
            has_visible_content = True

    merged_ranges = []
    for merged_range in root.findall("main:mergeCells/main:mergeCell", XML_NS):
        range_reference = merged_range.attrib.get("ref", "")
        if ":" not in range_reference:
            continue
        start_reference, end_reference = range_reference.split(":", 1)
        start_row, start_column = _cell_indexes(start_reference)
        end_row, end_column = _cell_indexes(end_reference)
        if end_row < row_start or start_row > row_end or end_column < column_start or start_column > column_end:
            continue
        merged_ranges.append(
            {
                "range": range_reference,
                "start_row": max(start_row, row_start),
                "end_row": min(end_row, row_end),
                "start_column": max(start_column, column_start),
                "end_column": min(end_column, column_end),
            }
        )

    row_dimensions = {
        int(row.attrib["r"]): row
        for row in root.findall(".//main:sheetData/main:row", XML_NS)
        if row.attrib.get("r", "").isdigit()
    }
    rows = []
    for index in range(row_start, row_end + 1):
        dimension = row_dimensions.get(index)
        rows.append(
            {
                "index": index,
                "height": float(dimension.attrib["ht"]) if dimension is not None and dimension.attrib.get("ht") else None,
                "hidden": dimension is not None and dimension.attrib.get("hidden") in ("1", "true"),
            }
        )

    column_dimensions = root.findall("main:cols/main:col", XML_NS)
    columns = []
    for index in range(column_start, column_end + 1):
        dimension = next(
            (
                item
                for item in column_dimensions
                if int(item.attrib.get("min", "0")) <= index <= int(item.attrib.get("max", "0"))
            ),
            None,
        )
        letters = ""
        remaining = index
        while remaining:
            remaining, remainder = divmod(remaining - 1, 26)
            letters = chr(65 + remainder) + letters
        columns.append(
            {
                "index": index,
                "letter": letters,
                "width": float(dimension.attrib["width"]) if dimension is not None and dimension.attrib.get("width") else None,
                "hidden": dimension is not None and dimension.attrib.get("hidden") in ("1", "true"),
            }
        )

    return {
        "worksheet_id": worksheet_id,
        "worksheet_name": worksheet_name,
        "row_start": row_start,
        "row_end": row_end,
        "column_start": column_start,
        "column_end": column_end,
        "total_rows": total_rows,
        "total_columns": total_columns,
        "is_empty": not has_visible_content,
        "is_bounded": row_end < total_rows or column_end < total_columns,
        "cells": cells,
        "merged_ranges": merged_ranges,
        "rows": rows,
        "columns": columns,
    }


def extract_original_workbook_layout(
    *,
    workbook_path: Path,
    worksheet_index: int,
    worksheet_id: str,
    row_start: int,
    row_limit: int,
    column_start: int,
    column_limit: int,
) -> dict[str, Any]:
    if workbook_path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Original workbook view currently supports XLSX files only")
    if not workbook_path.exists():
        raise HTTPException(status_code=404, detail="Original workbook file is missing")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise HTTPException(
            status_code=500,
            detail="Original workbook view requires the openpyxl backend dependency",
        ) from error

    bounded_row_limit = min(row_limit, MAX_ORIGINAL_LAYOUT_ROWS)
    bounded_column_limit = min(column_limit, MAX_ORIGINAL_LAYOUT_COLUMNS)

    try:
        workbook = load_workbook(
            filename=workbook_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception:
        return _extract_layout_from_ooxml(
            workbook_path=workbook_path,
            worksheet_index=worksheet_index,
            worksheet_id=worksheet_id,
            row_start=row_start,
            row_limit=bounded_row_limit,
            column_start=column_start,
            column_limit=bounded_column_limit,
        )

    try:
        if worksheet_index < 0 or worksheet_index >= len(workbook.worksheets):
            raise HTTPException(status_code=404, detail="Original workbook worksheet is missing")

        worksheet = workbook.worksheets[worksheet_index]
        total_rows = max(worksheet.max_row, 1)
        total_columns = max(worksheet.max_column, 1)
        row_end = min(total_rows, row_start + bounded_row_limit - 1)
        column_end = min(total_columns, column_start + bounded_column_limit - 1)

        cells: list[dict[str, Any]] = []
        has_visible_content = False
        for row in worksheet.iter_rows(
            min_row=row_start,
            max_row=row_end,
            min_col=column_start,
            max_col=column_end,
        ):
            for cell in row:
                display_value = _display_value(cell.value)
                has_style = bool(cell.has_style)
                if display_value or has_style:
                    cells.append(
                        {
                            "row": cell.row,
                            "column": cell.column,
                            "coordinate": cell.coordinate,
                            "display_value": display_value,
                            "is_formula": isinstance(cell.value, str) and cell.value.startswith("="),
                            "style": _cell_style(cell),
                        }
                    )
                if display_value:
                    has_visible_content = True

        merged_ranges = []
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.max_row < row_start
                or merged_range.min_row > row_end
                or merged_range.max_col < column_start
                or merged_range.min_col > column_end
            ):
                continue
            merged_ranges.append(
                {
                    "range": str(merged_range),
                    "start_row": max(merged_range.min_row, row_start),
                    "end_row": min(merged_range.max_row, row_end),
                    "start_column": max(merged_range.min_col, column_start),
                    "end_column": min(merged_range.max_col, column_end),
                }
            )

        rows = [
            {
                "index": index,
                "height": worksheet.row_dimensions[index].height,
                "hidden": bool(worksheet.row_dimensions[index].hidden),
            }
            for index in range(row_start, row_end + 1)
        ]
        columns = []
        for index in range(column_start, column_end + 1):
            letter = get_column_letter(index)
            dimension = worksheet.column_dimensions[letter]
            columns.append(
                {
                    "index": index,
                    "letter": letter,
                    "width": dimension.width,
                    "hidden": bool(dimension.hidden),
                }
            )

        return {
            "worksheet_id": worksheet_id,
            "worksheet_name": worksheet.title,
            "row_start": row_start,
            "row_end": row_end,
            "column_start": column_start,
            "column_end": column_end,
            "total_rows": total_rows,
            "total_columns": total_columns,
            "is_empty": not has_visible_content,
            "is_bounded": row_end < total_rows or column_end < total_columns,
            "cells": cells,
            "merged_ranges": merged_ranges,
            "rows": rows,
            "columns": columns,
        }
    finally:
        workbook.close()
